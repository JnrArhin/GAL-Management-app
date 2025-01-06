from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
import os
import openpyxl
import time
from datetime import datetime, date, timedelta
import openai

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize SQLAlchemy with the Flask app
db = SQLAlchemy()
db.init_app(app)

# Initialize Flask-Migrate
migrate = Migrate(app, db)

# Path to your Excel workbook
EXCEL_FILE_PATH = os.path.expanduser("C:\\Users\\USER\\Desktop\\GAL MM\\static\\mining_data.xlsx")

# Add your OpenAI API key
openai.api_key = 'your-api-key-here'

# User model for authentication
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    status = db.Column(db.String(50), default='pending')
    is_admin = db.Column(db.Boolean, default=False)
    last_login = db.Column(db.DateTime)

# Drop and recreate all tables
with app.app_context():
    db.drop_all()  # This will delete all existing data
    db.create_all()
    # Create admin user
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin_user = User(username='admin', password='goldstonedbapp', status='approved', is_admin=True)
        db.session.add(admin_user)
        db.session.commit()
        
class AIEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(100), nullable=False)
    status = db.Column(db.String(50), default='active')
    skills = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_active = db.Column(db.DateTime, default=datetime.utcnow)

# Function to save data to the specified sheet
def save_to_excel(sheet_name, data):
    while True:
        try:
            if os.path.exists(EXCEL_FILE_PATH):
                workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            else:
                workbook = openpyxl.Workbook()

            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
                headers = list(data.keys())
                sheet.append(headers)
            else:
                sheet = workbook[sheet_name]

            sheet.append(list(data.values()))
            workbook.save(EXCEL_FILE_PATH)
            print(f"Data saved to {sheet_name}: {data}")
            print(f"Data saved to file: {EXCEL_FILE_PATH}")
            break
        except PermissionError:
            print("Workbook is open. Waiting to try again...")
            time.sleep(5)
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            break

# Function to check for duplicates
def check_for_duplicates(sheet_name, date, shift, equipment_id=None):
    if os.path.exists(EXCEL_FILE_PATH):
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0] == date and row[1] == shift:  # Check only date and shift
                    return True
    return False

def check_equipment_stats_duplicate(date, shift, equipment_id):
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            if 'Equipment Statistics' in workbook.sheetnames:
                sheet = workbook['Equipment Statistics']
                
                # Find the column indices for date, shift, and equipment ID
                headers = [cell.value for cell in sheet[1]]
                date_col = headers.index('Mining Date') + 1
                shift_col = headers.index('Mining Shift') + 1
                equip_col = headers.index('Equipment ID') + 1
                
                # Check each row for matches
                for row in sheet.iter_rows(min_row=2):  # Skip header row
                    row_date = row[date_col-1].value
                    row_shift = row[shift_col-1].value
                    row_equipment = row[equip_col-1].value
                    
                    if (row_date == date and 
                        row_shift == shift and 
                        row_equipment == equipment_id):
                        return True
                        
            workbook.close()
    except Exception as e:
        print(f"Error checking for duplicates: {e}")
    
    return False

def check_incident_duplicate(incident_date, incident_time, involved_person, incident_location):
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            if 'Incident Reports' in workbook.sheetnames:
                sheet = workbook['Incident Reports']
                
                # Find the column indices
                headers = [cell.value for cell in sheet[1]]
                date_col = headers.index('incident-date') + 1
                time_col = headers.index('incident-time') + 1
                person_col = headers.index('involved-person') + 1
                location_col = headers.index('incident-location') + 1
                
                # Check each row for matches
                for row in sheet.iter_rows(min_row=2):  # Skip header row
                    row_date = str(row[date_col-1].value)
                    row_time = str(row[time_col-1].value)
                    row_person = str(row[person_col-1].value)
                    row_location = str(row[location_col-1].value)
                    
                    if (row_date == incident_date and 
                        row_time == incident_time and 
                        row_person == involved_person and 
                        row_location == incident_location):
                        return True
                        
            workbook.close()
    except Exception as e:
        print(f"Error checking for incident duplicates: {e}")
    
    return False

def check_hazard_duplicate(hazard_date, hazard_time, hazard_location, reported_by):
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            if 'Hazard Reports' in workbook.sheetnames:
                sheet = workbook['Hazard Reports']
                
                # Find the column indices
                headers = [cell.value for cell in sheet[1]]
                date_col = headers.index('hazard-date') + 1
                time_col = headers.index('hazard-time') + 1
                location_col = headers.index('hazard-location') + 1
                reporter_col = headers.index('reported-by') + 1
                
                # Check each row for matches
                for row in sheet.iter_rows(min_row=2):  # Skip header row
                    row_date = str(row[date_col-1].value)
                    row_time = str(row[time_col-1].value)
                    row_location = str(row[location_col-1].value)
                    row_reporter = str(row[reporter_col-1].value)
                    
                    if (row_date == hazard_date and 
                        row_time == hazard_time and 
                        row_location == hazard_location and 
                        row_reporter == reported_by):
                        return True
                        
            workbook.close()
    except Exception as e:
        print(f"Error checking for hazard duplicates: {e}")
    
    return False

@app.route('/excel/<path:filename>')
def serve_excel(filename):
    return send_from_directory('static', filename)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Check if the username already exists
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            return "Username already exists. Please choose a different one.", 400  # Bad request

        new_user = User(username=username, password=password, status='pending')
        db.session.add(new_user)
        db.session.commit()
        return redirect(url_for('login'))  # Redirect to login after registration
    return render_template('register.html')  # Render the registration page

@app.route('/admin/registrations')
def view_registrations():
    if not is_admin():  # Ensure only admins can access this
        return redirect(url_for('login'))
    pending_users = User.query.filter_by(status='pending').all()
    return render_template('admin_registrations.html', users=pending_users)

@app.route('/admin/approve/<int:user_id>')
def approve_user(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    
    user = User.query.get(user_id)
    if user:
        user.status = 'approved'  # Set status to approved
        db.session.commit()
        flash('User approved successfully!', 'success')
    return redirect(url_for('view_registrations'))

@app.route('/admin/reject/<int:user_id>')
def reject_user(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    
    user = User.query.get(user_id)
    if user:
        user.status = 'rejected'  # Set status to rejected
        db.session.commit()
        flash('User rejected successfully!', 'danger')
    return redirect(url_for('view_registrations'))

def is_admin():
    # Implement your logic to check if the current user is an admin
    return 'admin' in session and session['admin']

@app.route('/')
def index():
    return redirect(url_for('login'))  # Redirect to the login page

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username, password=password).first()
        
        if user:
            # Check if user is admin or already approved
            if user.is_admin or user.status == 'approved':
                session['user_id'] = user.id
                session['admin'] = user.is_admin
                user.last_login = datetime.utcnow()
                db.session.commit()
                print(f"User {username} logged in successfully.")
                return redirect(url_for('dashboard'))
            else:
                return "Your account is pending approval", 401
        else:
            return "Invalid credentials", 401
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        print("User not authenticated, redirecting to login.")
        return redirect(url_for('login'))  # Redirect to login if not authenticated
    print("Rendering dashboard.")
    return render_template('index3.html')  # Render the dashboard page

@app.route('/submit_mining_material_data', methods=['POST'])
def submit_mining_material_data():
    mining_date = request.form['mining-date']
    mining_shift = request.form['mining-shift']
    user_action = request.form.get('user_action', 'none')  # "continue", "overwrite", or "cancel"

    # Check for duplicates
    is_duplicate = check_for_duplicates('Mining Materials', mining_date, mining_shift)
    if is_duplicate and user_action == 'none':
        return jsonify({
            "error": "Duplicate entry found.",
            "options": ["continue", "overwrite", "cancel"]
        }), 409  # Conflict status code

    data = {
        'Date': mining_date,
        'Shift': mining_shift,
        'Total Ore Truck Count': request.form ['ore-truck-count'],
        'Total Waste Truck Count': request.form ['waste-truck-count']
    }

    try:
        save_to_excel('Mining Materials', data)
        return jsonify({"message": "Mining materials submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_equipment_stats_data', methods=['POST'])
def submit_equipment_stats_data():
    mining_date = request.form['mining-date']
    mining_shift = request.form['mining-shift']
    equipment_id = request.form['equipment-id']
    user_action = request.form.get('user_action', 'none')

    # Use the new duplicate check function
    is_duplicate = check_equipment_stats_duplicate(mining_date, mining_shift, equipment_id)
    
    if is_duplicate and user_action == 'none':
        return jsonify({
            "error": "Duplicate entry found.",
            "options": ["continue", "overwrite", "cancel"]
        }), 409

    if is_duplicate and user_action == 'cancel':
        return jsonify({"message": "Action cancelled by the user."}), 200

    if is_duplicate and user_action == 'continue':
        # Proceed without overwriting
        return jsonify({"message": "Data submission continued without changes."}), 200

    # If no duplicate or user chooses to overwrite
    data = {
        'Mining Date': mining_date,
        'Mining Shift': mining_shift,
        'Equipment ID': equipment_id,
        'Start Hour Meter': request.form['start-hour-meter'],
        'End Hour Meter': request.form['end-hour-meter'],
        'Equipment Run Hours': request.form['equipment-run-hours'],
        'Fuel Recieved': request.form['fuel-recieved'],
        'Mining Operation Start Time': request.form['mining-operations-start-time'],
        'Mining Operation End Time': request.form['mining-operations-end-time'],
        'Total Production Hours': request.form['total-production-hours'],
        'Total Downtime Hours': request.form['total-downtime-hours'],
        'Total Tramming Hours': request.form['total-tramming-hours'],
        'Operational GSA Hours': request.form['operational-gsa-hours'],
        'Other GSA Hours': request.form['other-gsa-hours'],
        'Equipment Standby': request.form['equipment-standby'],
        'Comments': request.form['comments'],
        'Operator Name': request.form['operator-name'],
        'Supervisor Name': request.form['Supervisor Name'],
    }

    try:
        save_to_excel('Equipment Statistics', data)
        return jsonify({"message": "Equipment stats submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_exploration_geology_metrics', methods=['POST'])
def submit_exploration_geology_metrics():
    field_selection = request.form['field-selection']
    supervisor = request.form['supervisor']

    data = {
        'Field Selection': field_selection,
        'Supervisor': supervisor
    }

    # Add specific data based on the selected field
    if field_selection == 'mapping':
        data.update({
            'Map Date': request.form['map-date'],
            'Map Shift': request.form['map-shift'],
            'Map Scale': request.form['map-scale'],
            'Traverse Distance': request.form['traverse-distance'],
            'Area Covered': request.form['area-covered'],
        })
    elif field_selection == 'geophysical':
        data.update({
            'Survey Date': request.form['survey-date'],
            'Survey Shift': request.form['survey-shift'],
            'Survey Method': request.form['survey-method'],
            'Survey Area': request.form['survey-area'],
            'Equipment Used': request.form['equipment-used'],
            'Anomalies Detected': request.form['anomalies-detected'],
            'Interpretation': request.form['interpretation'],
        })
    elif field_selection == 'geochemical':
        data.update({
            'Geochem Date': request.form['geochem-date'],
            'Geochem Shift': request.form['geochem-shift'],
            'Sample ID Range': request.form['sample-id-range'],
            'Number of Samples': request.form['num-samples'],
            'Sampling Geologist': request.form['sampling-geologist'],
            'Sampling Method': request.form['sampling-method'],
            'Laboratory': request.form['laboratory'],
        })
    elif field_selection == 'trenching':
        data.update({
            'Trench Date': request.form['trench-date'],
            'Trench Shift': request.form['trench-shift'],
            'Trench ID': request.form['trench-id'],
            'Trench Length': request.form['trench-length'],
            'Trench Width': request.form['trench-width'],
            'Sample Results': request.form['sample-results'],
            'Geological Description': request.form['geological-description'],
            'Channel Sampling': request.form['channel-sampling'],
        })
    elif field_selection == 'drilling':
        data.update({
            'Drilling Date': request.form['drilling-date'],
            'Drilling Shift': request.form['drilling-shift'],
            'Drill Hole ID': request.form['drill-hole-id'],
            'Depth': request.form['depth'],
            'Core Recovery': request.form['core-recovery'],
            'Lithology': request.form['lithology'],
            'Structural Analysis': request.form['structural-analysis'],
            'Downhole Logs': request.form['downhole-logs'],
            'Sampling Technique': request.form['sampling-technique'],
        })

    try:
        save_to_excel('Exploration Geology Metrics', data)
        return jsonify({"message": "Exploration geology metrics submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Hazard Reports Section
@app.route('/submit_hazard_report_data', methods=['POST'])
def submit_hazard_report_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Hazard Reports', data)
        return jsonify({"message": "Hazard report data saved successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Maintenance Data Section
@app.route('/submit_maintenance_data', methods=['POST'])
def submit_maintenance_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Maintenance Data', data)
        return jsonify({"message": "Maintenance data saved successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_crushing_data', methods=['POST'])
def submit_crushing_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Crushing Data', data)
        return jsonify({"message": "Crushing data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_solution_management_data', methods=['POST'])
def submit_solution_management_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Solution Management Data', data)
        return jsonify({"message": "Solution management data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_geophysics_data', methods=['POST'])
def submit_geophysics_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Geophysical Data', data)
        return jsonify({"message": "Geophysical data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_geochemical_data', methods=['POST'])
def submit_geochemical_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Geochemical Data', data)
        return jsonify({"message": "Geochemical data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_trenching_data', methods=['POST'])
def submit_trenching_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Trenching Data', data)
        return jsonify({"message": "Trenching data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_drilling_data', methods=['POST'])
def submit_drilling_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Drilling Data', data)
        return jsonify({"message": "Drilling data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_incident_report_data', methods=['POST'])
def submit_incident_report_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Incident Reports', data)
        return jsonify({"message": "Incident report submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_monitoring_data', methods=['POST'])
def submit_monitoring_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Air and Noise Monitoring', data)
        return jsonify({"message": "Monitoring data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_water_sample_data', methods=['POST'])
def submit_water_sample_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Water Samples', data)
        return jsonify({"message": "Water sample data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_inspection_data', methods=['POST'])
def submit_inspection_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Site Inspections', data)
        return jsonify({"message": "Inspection data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_complaint_data', methods=['POST'])
def submit_complaint_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Complaints', data)
        return jsonify({"message": "Complaint submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_requests_data', methods=['POST'])
def submit_requests_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Requests', data)
        return jsonify({"message": "Request submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_stakeholder_engagement_data', methods=['POST'])
def submit_stakeholder_engagement_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Stakeholder Engagement', data)
        return jsonify({"message": "Stakeholder engagement data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_assessment_data', methods=['POST'])
def submit_assessment_data():
    data = request.form.to_dict()
    try:
        save_to_excel('Land and Crop Assessments', data)
        return jsonify({"message": "Land and crop assessment data submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get_kpis', methods=['GET'])
def get_kpis():
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            
            # Initialize KPI data structure
            kpi_data = {
                'excavator': calculate_equipment_kpis('EX'),
                'dozer': calculate_equipment_kpis('DZ'),
                'grader': calculate_equipment_kpis('GR'),
                'waterkart': calculate_equipment_kpis('WK'),
                'towerlight': calculate_equipment_kpis('TL'),
                'pump': calculate_equipment_kpis('WP')
            }
            
            return jsonify(kpi_data), 200
    except Exception as e:
        print(f"Error getting KPIs: {e}")
        return jsonify({"error": str(e)}), 500

def calculate_equipment_kpis(equipment_code):
    """Calculate KPIs for specific equipment type"""
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            sheet = workbook['Equipment Statistics']
            
            # Get today's date
            today = datetime.now().date()
            
            # Initialize counters
            total_production_hours = 0
            total_downtime_hours = 0
            entries_count = 0
            
            # Process each row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[2] and str(row[2]).upper().startswith(equipment_code):  # Equipment ID check
                    entry_date = datetime.strptime(row[0], '%Y-%m-%d').date()
                    if entry_date == today:
                        production_hours = float(row[9]) if row[9] else 0  # Total Production Hours
                        downtime_hours = float(row[10]) if row[10] else 0  # Total Downtime Hours
                        
                        total_production_hours += production_hours
                        total_downtime_hours += downtime_hours
                        entries_count += 1
            
            # Calculate KPIs
            scheduled_hours = entries_count * 12  # 12 hours per shift
            
            if scheduled_hours > 0:
                availability = ((scheduled_hours - total_downtime_hours) / scheduled_hours) * 100
                utilization = (total_production_hours / scheduled_hours) * 100
            else:
                availability = 0
                utilization = 0
                
            return {
                'availability': f"{availability:.1f}%",
                'utilization': f"{utilization:.1f}%",
                'production_hours': total_production_hours,
                'downtime_hours': total_downtime_hours
            }
            
    except Exception as e:
        print(f"Error calculating KPIs for {equipment_code}: {e}")
        return {
            'availability': "0%",
            'utilization': "0%",
            'production_hours': 0,
            'downtime_hours': 0
        }

@app.route('/submit_incident_report', methods=['POST'])
def submit_incident_report():
    incident_date = request.form['incident-date']
    incident_time = request.form['incident-time']
    involved_person = request.form['involved-person']
    incident_location = request.form['incident-location']
    user_action = request.form.get('user_action', 'none')

    # Check for duplicates
    is_duplicate = check_incident_duplicate(incident_date, incident_time, involved_person, incident_location)
    
    if is_duplicate and user_action == 'none':
        return jsonify({
            "error": "Duplicate incident report found.",
            "options": ["continue", "overwrite", "cancel"]
        }), 409

    if is_duplicate and user_action == 'cancel':
        return jsonify({"message": "Action cancelled by user."}), 200

    data = request.form.to_dict()
    
    try:
        save_to_excel('Incident Reports', data)
        return jsonify({"message": "Incident report submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_hazard_report', methods=['POST'])
def submit_hazard_report():
    hazard_date = request.form['hazard-date']
    hazard_time = request.form['hazard-time']
    hazard_location = request.form['hazard-location']
    reported_by = request.form['reported-by']
    user_action = request.form.get('user_action', 'none')

    # Check for duplicates
    is_duplicate = check_hazard_duplicate(hazard_date, hazard_time, hazard_location, reported_by)
    
    if is_duplicate and user_action == 'none':
        return jsonify({
            "error": "Duplicate hazard report found.",
            "options": ["continue", "overwrite", "cancel"]
        }), 409

    if is_duplicate and user_action == 'cancel':
        return jsonify({"message": "Action cancelled by user."}), 200

    data = request.form.to_dict()
    
    try:
        save_to_excel('Hazard Reports', data)
        return jsonify({"message": "Hazard report submitted successfully!"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get_injuries_history')
def get_injuries_history():
    try:
        # Get current date
        current_date = datetime.now()
        
        # Calculate date ranges
        week_ago = current_date - timedelta(days=7)
        month_ago = current_date - timedelta(days=30)
        year_ago = current_date - timedelta(days=365)

        # Read Excel file
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook['Incident Reports']
        
        # Initialize data containers
        daily_data = []
        weekly_data = []
        monthly_data = []
        yearly_data = []

        # Process data
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            incident_date = row[0]  # Assuming date is in first column
            if isinstance(incident_date, datetime):
                if (current_date - incident_date).days <= 7:
                    daily_data.append(row)
                if (current_date - incident_date).days <= 30:
                    weekly_data.append(row)
                if (current_date - incident_date).days <= 90:
                    monthly_data.append(row)
                if (current_date - incident_date).days <= 365:
                    yearly_data.append(row)

        return jsonify({
            'current': len(daily_data),
            'historical': [len(d) for d in daily_data],
            'weeklyTotal': len(weekly_data),
            'weeklyHistory': [len(d) for d in weekly_data],
            'monthlyTotal': len(monthly_data),
            'monthlyHistory': [len(d) for d in monthly_data],
            'yearlyTotal': len(yearly_data),
            'yearlyHistory': [len(d) for d in yearly_data]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/users')
def manage_users():
    if not is_admin():
        return redirect(url_for('login'))
    users = User.query.filter(User.username != 'admin').all()
    return render_template('admin_user_management.html', users=users)

@app.route('/admin/block/<int:user_id>')
def block_user(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    user = User.query.get(user_id)
    if user and user.username != 'admin':
        user.status = 'blocked'
        db.session.commit()
        flash(f'User {user.username} has been blocked')
    return redirect(url_for('manage_users'))

@app.route('/admin/unblock/<int:user_id>')
def unblock_user(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    user = User.query.get(user_id)
    if user:
        user.status = 'active'
        db.session.commit()
        flash(f'User {user.username} has been unblocked')
    return redirect(url_for('manage_users'))

@app.route('/admin/remove/<int:user_id>')
def remove_user(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    user = User.query.get(user_id)
    if user and user.username != 'admin':
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.username} has been removed')
    return redirect(url_for('manage_users'))

@app.route('/get_ore_trend')
def get_ore_trend():
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            sheet = workbook['Mining Materials']
            
            # Get today's date
            today = datetime.now().date()
            
            # Initialize data structure for last 7 days
            daily_data = []
            trend_data = {
                'dates': [],
                'ore_counts': [],
                'trend_direction': '',
                'percentage_change': 0
            }
            
            # Collect data for the last 7 days
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    entry_date = datetime.strptime(row[0], '%Y-%m-%d').date()
                    if (today - entry_date).days <= 7:  # Last 7 days
                        daily_data.append({
                            'date': entry_date,
                            'ore_count': int(row[2]) if row[2] else 0  # Total Ore Truck Count
                        })
                except (ValueError, TypeError):
                    continue
            
            # Sort data by date
            daily_data.sort(key=lambda x: x['date'])
            
            # Calculate trend
            if len(daily_data) >= 2:
                first_value = daily_data[0]['ore_count']
                last_value = daily_data[-1]['ore_count']
                
                # Calculate percentage change
                if first_value > 0:
                    percentage_change = ((last_value - first_value) / first_value) * 100
                else:
                    percentage_change = 0
                
                # Determine trend direction
                if percentage_change > 0:
                    trend_direction = 'increasing'
                elif percentage_change < 0:
                    trend_direction = 'decreasing'
                else:
                    trend_direction = 'stable'
                
                # Prepare data for chart
                trend_data['dates'] = [d['date'].strftime('%Y-%m-%d') for d in daily_data]
                trend_data['ore_counts'] = [d['ore_count'] for d in daily_data]
                trend_data['trend_direction'] = trend_direction
                trend_data['percentage_change'] = round(percentage_change, 2)
            
            return jsonify(trend_data)
            
    except Exception as e:
        print(f"Error getting ore trend: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    with app.app_context():  # Ensure the application context is set
        db.create_all()  # Create tables if they don't exist
    app.run(host='0.0.0.0', debug=True)  # Allow access from other computers